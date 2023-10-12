# -*-coding:Latin-1 -*
# Les réponses aux input sont considérés négatives pour toute phrase commençant
# par un n et positives pour celles commençant par une autre lettre

import configparser
import os
import string
import sys

from geopy import geocoders
from geopy.distance import geodesic
from openpyxl import load_workbook  # to open Excel formats.

gn = geocoders.GeoNames(username='oaumont')  # 'oaumont'

# Conditions for planes, short-mid-long distance (in km)
d_short_distance_plane = 1000
d_medium_distance_plane = 3500

# Emissions factors for plane (kgCO2e / km)
offset_plane_km = 95  # an offset to add to the geodesic distance for plane.
FE_short_distance_plane = 0.258
FE_medium_distance_plane = 0.187
FE_long_distance_plane = 0.152

# Emission factor for cars (general) (kgCO2e / km)
FE_cars = 0.233
factor_distance_car = 1.3  # approx. : multiply the geodesic distance to take into account the path is not straight.

# Emissions factors for train (kgCO2e / km)
factor_distance_train = 1.2  # approx. : multiply the geodesic distance to take into account the path is not straight.
d_TER_TGV = 200  # limite distance to take TER or TGV emissions
FE_french_TER = 0.018
FE_french_TGV = 0.003
FE_half_french_trains = 0.016  # if the departure is in France but not the destination (vice-versa)
FE_european_trains = 0.037  # for trains between two cities outside France

# Seuil/threshold of distance (km) used only if a row is missing the type of transport.
Dseuil = 700

# data files
my_config_file_path = 'Data/config_file_LMD_CNRS_2022.cfg'
my_data_file_path = f"Data/MIS_2022_extrait.xlsx"  # Supported formats are: .xlsx,.xlsm,.xltx,.xltm

# check if the config file is found
if not os.path.exists(my_config_file_path):
    print('no config file found')
    sys.exit()

##### MAIN PROGRAM ######

# read the Excel listing
wb = load_workbook(filename=my_data_file_path)
res = wb.sheetnames  # the name of the excel sheet, from the, example [Sheet1]
n_res = len(wb.sheetnames)  # number of sheets

# read the configuration file
config = configparser.RawConfigParser()
config.read(my_config_file_path)


# two useful fonctions needed thereafter
def rchop(thestring, ending):
    """ Julie = crops the string to remove the ending part if it is present."""
    if thestring.endswith(ending):
        return thestring[:-len(ending)]
    return thestring


def compute_letter_next_columns(nlist: int, add: int):
    """ Computes the letter corresponding to the last column plus the number add.
    It has to work in modulo 26 since 27 numbers in alphabet."""
    n = nlist
    aa = (n + add) // 26

    bb = (n + add) % 26
    c = []
    if aa > 0:
        d = "a"
        for i in range(aa):
            c = c + list(d)
    f = list(string.ascii_lowercase[bb])
    c = c + f
    g = ''.join(c)
    return g


# Search of the departure and arrival columns and compute the CO2 emissions :
ws_sheet = res[0]  # the name of the first sheet
ws = wb[ws_sheet]  # the fist excel sheet

# takes the columns head names
column_list = []
for cell in ws[1]:  # what does this do ?
    if cell.value:
        column_list.append(int(cell.column))
n_column_list = len(column_list) - 1

# takes the letters of the useful columns, from the config file, for destination city, etc.
config_dict = dict(config.items(res[0]))  # res[i_sheet] is the excel sheet number i_sheet+1.
key_list = list(config_dict.keys())

# extract the column that are useful for us, from the letters in the config file.
col_ar = ws[config_dict['aller_retour']]  # columns of the answer to 'aller-retour', oui, non
col_depart = ws[config_dict['depart_column']]
col_land_departure = ws[config_dict['land_depart_column']]
col_arrival = ws[config_dict['arrival_column']]  # Liste des villes d'arrivées
col_land_arrival = ws[config_dict['land_arrival_column']]
col_type = ws[config_dict['t_type_column']]

# names of the type of transport in this excel sheet
t_type_car = config_dict['t_type_car']
t_type_train = config_dict['t_type_train']
t_type_air = config_dict['t_type_air']

# Create new columns to add the values of distance, CO2 and comment
col_distance = ws[compute_letter_next_columns(n_column_list, 1)]
col_CO2 = ws[compute_letter_next_columns(n_column_list, 2)]
col_comment = ws[compute_letter_next_columns(n_column_list, 3)]
# Add headings for them
col_CO2[0].value = "Emissions (kgCO2e)"
col_comment[0].value = "Validite"
col_distance[0].value = "Distance (km)"

# compute the distances and CO2 emissions

for i_rows in range(1, len(col_depart)):  # omit the first line, the column titles

    print(f"working on row number {i_rows + 1}/{len(col_depart)}")

    # compute the localisation
    if col_land_departure[i_rows].value is not None:
        departure = col_depart[i_rows].value + " (" + col_land_departure[i_rows].value + ")"
    else:
        departure = col_depart[i_rows].value
    if col_land_arrival[i_rows].value is not None:
        arrival = col_arrival[i_rows].value + " (" + col_land_arrival[i_rows].value + ")"
    else:
        arrival = col_arrival[i_rows].value

    # Using geocode, find the full locations (with latitude and longitude) of the cities of departure and arrival.
    locations_found = True
    try:
        location_d = gn.geocode(departure, timeout=30)
        location_a = gn.geocode(arrival, timeout=30)
        if location_d is None:
            raise ValueError(
                f'Departure location not found, at row {i_rows} for {departure}')  # GeocoderTimedOut, not defined...

        if location_a is None:
            raise ValueError(f'Arrival location not found, at row {i_rows} for {arrival}')
    except ValueError as e:
        # print("Error: geocode failed with message %s"%(e.message))
        locations_found = False

    # now compute the geodesic distance in km, then the related CO2 emissions

    if locations_found is True:
        dist_km: float = round(geodesic((location_d.latitude, location_d.longitude),
                                        (location_a.latitude, location_a.longitude)).km)

        transport_type = col_type[i_rows].value
        # print(transport_type)

        # modify the excel if information is missing
        if transport_type is None:  # if no given transport
            transport_type = "No given transport"
            print(transport_type)

        # We can now apply the plane emission factors, if at least a plane is mentionned (greater emissions than train)
        t_type_main = 0
        if t_type_air in transport_type:
            t_type_main = t_type_air
            if float(dist_km) < d_short_distance_plane:
                emp_co2 = float(dist_km + offset_plane_km) * FE_short_distance_plane
            elif float(dist_km) < d_medium_distance_plane:
                emp_co2 = float(dist_km + offset_plane_km) * FE_medium_distance_plane
            else:  # long-distance plane
                emp_co2 = float(dist_km + offset_plane_km) * FE_long_distance_plane


        # Apply the train emissions, for the train traject (on which no plane is taken at all)
        elif t_type_train in transport_type:
            t_type_main = t_type_train
            # For train emission factor, we distinguish if both cities are in France or not, to use SNCF's emission factors or European ones.
            if (location_d.raw['countryCode'] == 'FR') and (location_a.raw['countryCode'] == 'FR'):
                if factor_distance_train * dist_km < d_TER_TGV:
                    # use TER emission factors
                    emp_co2 = factor_distance_train * float(dist_km) * FE_french_TER
                else:
                    emp_co2 = factor_distance_train * float(dist_km) * FE_french_TGV
            elif (location_d.raw['countryCode'] == 'FR') or (location_a.raw['countryCode'] == 'FR'):
                emp_co2 = factor_distance_train * float(dist_km) * FE_half_french_trains
            else:
                emp_co2 = factor_distance_train * float(dist_km) * FE_european_trains

        # Apply cars emission, when it's the main transport.
        elif t_type_car.split(',')[0] in transport_type or t_type_car.split(',')[1] in transport_type:
            t_type_main = t_type_car
            emp_co2 = float(dist_km * factor_distance_car) * FE_cars

        # Finally the rest, rows with missing information, or rows with neither train or plane or car, but only bus, metro etc. We assume a type of transport according to the distance.
        else:
            print('Transport type based on distance')
            if dist_km < Dseuil:  # we consider it is train
                t_type_main = t_type_train
                col_type[i_rows].value = t_type_train
            else:  # we consider that a plane was used
                t_type_main = t_type_air
                col_type[i_rows].value = t_type_air

        print(f"One-way emissions from {departure} to {arrival} ({round(dist_km)} km) by {t_type_main} is {round(emp_co2,1)} kg C02e")

    aller_retour = col_ar[i_rows].value
    if aller_retour in ['oui', 'OUI', "Yes", "YES"]:
        #  it is an 'aller-retour', multiply distance and CO2 by 2
        dist_km *= 2
        emp_co2 *= 2

    # write those values in the excel file
    if locations_found:
        col_CO2[i_rows].value = round(emp_co2, 1)
        col_comment[i_rows].value = 'OK'
        col_distance[i_rows].value = dist_km
        # add the country code, instead of the country name
        col_land_departure[i_rows].value = location_d.raw['countryCode']
        col_land_arrival[i_rows].value = location_a.raw['countryCode']
        
    else:
        col_comment[i_rows].value = 'Location not found'
        col_distance[i_rows].value = float(0)
        col_CO2[i_rows].value = float(0)

# saving the data as a new file
results_file_path = rchop(my_data_file_path, '.xlsx') + '_CO2.xlsx'
wb.save(filename=results_file_path)
print(f"Excel resulting file saved at this adress: {results_file_path}")
print('Program Empr_missions.py completed !')
